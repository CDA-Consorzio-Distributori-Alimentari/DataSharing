from io import BytesIO
from pathlib import Path

import openpyxl

class ExcelManager:
    def _build_workbook(self, sheet_name, data):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        for row in data:
            sheet.append(row)

        return workbook

    def build_excel_content(self, sheet_name, data):
        workbook = self._build_workbook(sheet_name, data)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def write_excel(self, filename, sheet_name, data):
        excel_content = self.build_excel_content(sheet_name, data)
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        with open(filename, 'wb') as file:
            file.write(excel_content)
        return excel_content

    def create_workbook(self, filename):
        workbook = openpyxl.Workbook()
        workbook.save(filename)

    def write_to_sheet(self, filename, sheet_name, data):
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        for row in data:
            sheet.append(row)

        workbook.save(filename)