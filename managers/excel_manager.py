import openpyxl

class ExcelManager:
    def create_workbook(self, filename):
        workbook = openpyxl.Workbook()
        workbook.save(filename)

    def write_to_sheet(self, filename, sheet_name, data):
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        for row in data:
            sheet.append(row)

        workbook.save(filename)